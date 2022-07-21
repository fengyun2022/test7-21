import openpyxl

class ReadWrite:
    def __init__(self,file,sheetname):
        self.file=file
        self.sheetname=sheetname
        self.wb = openpyxl.load_workbook(self.file)
        self.table = self.wb[self.sheetname]
        self.rows = self.table.max_row
        self.ncols = self.table.max_column

    def read_data(self):
        list2=[]
        for i in range(2,self.rows+1):
            list1 = []
            for j in range(1,self.ncols+1):
                value=self.table.cell(i,j).value
                list1.append(value)
            list2.append(list1)
        return list2


    def write_data(self,*values):
        for row in range(self.rows+1,self.rows+2):
            for col in range(1,self.ncols+1):
                self.table.cell(row,col).value=values[col-1]
        self.wb.save(self.file)
